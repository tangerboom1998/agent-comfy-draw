#!/usr/bin/env python3
"""Export ALL pretags categories to a single multi-sheet Excel file.

Usage:
    python scripts/pretags_export_all.py [output_path]

Default output: tmp/pretags_full_export.xlsx (relative to CWD)

Each category gets its own sheet with:
- Frozen header row + auto-filter
- Lora=1 rows highlighted green (#E2EFDA)
- Category-specific columns (人物 has name/外貌/服装, 画风 has 画风描述, etc.)
"""

import json, os, sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PRETAGS = os.path.join(os.path.dirname(__file__), "..", "..", "pretags.json")
DEFAULT_OUT = "tmp/pretags_full_export.xlsx"

def safe_str(v):
    if v is None or v == 0:
        return ""
    return str(v)

def col_letter(i):
    if i <= 26:
        return chr(64 + i)
    return chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)

def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    with open(PRETAGS, "r") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    lora_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # (sheet_title, data_key, headers, fields, widths)
    # sheet_title: Excel 工作表显示名（中文）
    # data_key: JSON 数据键名（英文，已迁移）
    categories = [
        ("人物", "characters",
         ["序号", "cname", "Source", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "name", "appearance", "clothing"],
         ["cname", "Source", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "name", "appearance", "clothing"],
         [6, 30, 16, 6, 40, 15, 12, 12, 50, 40, 60, 60]),
        ("画风", "style",
         ["序号", "cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag", "画风描述"],
         ["cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag", "画风描述"],
         [6, 30, 6, 40, 15, 12, 12, 50, 50, 80]),
        ("服装", "clothing",
         ["序号", "cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag"],
         ["cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag"],
         [6, 30, 6, 40, 15, 12, 12, 50, 50]),
        ("动作", "action",
         ["序号", "cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag"],
         ["cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag"],
         [6, 30, 6, 40, 15, 12, 12, 50, 50]),
        ("镜头", "shot",
         ["序号", "cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag"],
         ["cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag"],
         [6, 30, 6, 40, 15, 12, 12, 50, 50]),
        ("场景", "scene",
         ["序号", "cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag"],
         ["cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag"],
         [6, 30, 6, 40, 15, 12, 12, 50, 50]),
        ("其他", "other",
         ["序号", "cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag", "d_group"],
         ["cname", "Lora", "model file name", "model hash", "unet weight", "clip weight", "link", "tag", "d_group"],
         [6, 30, 6, 40, 15, 12, 12, 50, 50, 20]),
    ]

    total = 0
    for sheet_name, cat_key, headers, fields, widths in categories:
        ws = wb.create_sheet(title=sheet_name)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        if cat_key == "characters":
            entries = data.get("characters", {})
        else:
            entries = data.get("categories", {}).get(cat_key, {})
        row = 2
        for cname in sorted(entries.keys()):
            val = entries[cname]
            if not isinstance(val, dict):
                continue
            ws.cell(row=row, column=1, value=row - 1).border = thin_border
            for i, field in enumerate(fields):
                cell = ws.cell(row=row, column=i + 2, value=safe_str(val.get(field, "")))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if val.get("Lora") == 1:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = lora_fill
            row += 1

        for i, w in enumerate(widths):
            ws.column_dimensions[col_letter(i + 1)].width = w
        ws.freeze_panes = "A2"
        last_row = row - 1
        if last_row > 1:
            ws.auto_filter.ref = f"A1:{col_letter(len(headers))}{last_row}"
        count = row - 2
        print(f"  {sheet_name}: {count} rows")
        total += count

    wb.save(out_path)
    size_kb = os.path.getsize(out_path) / 1024
    print(f"\nTotal: {total} entries -> {out_path} ({size_kb:.0f} KB)")

if __name__ == "__main__":
    main()
