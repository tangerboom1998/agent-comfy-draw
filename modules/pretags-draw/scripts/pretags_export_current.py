#!/usr/bin/env python3
"""Export pretags.json to multi-sheet Excel (current schema).

New structure (2026-05):
- characters: 19K+ entries with cname, source, appearance, clothing, lora info
- categories: 画风/服装/动作/镜头/场景/其他
- series: 2K+ entries with name, count, characters

Usage:
    python scripts/pretags_export_current.py [output_path]
"""

import json, os, sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PRETAGS = os.path.join(os.path.dirname(__file__), "..", "..", "pretags.json")
DEFAULT_OUT = "tmp/pretags_full_export.xlsx"

def col_letter(i):
    if i <= 26:
        return chr(64 + i)
    return chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)

def make_ws(wb, title, headers, widths, rows_fn):
    """Create a formatted worksheet."""
    ws = wb.create_sheet(title=title)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    lora_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    ncols = len(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    rows_fn(ws, thin_border, lora_fill)
    for i, w in enumerate(widths):
        ws.column_dimensions[col_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    last_row = ws.max_row
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{col_letter(ncols)}{last_row}"
    return ws

def safe_str(v):
    if v is None or v == 0:
        return ""
    return str(v)

def main(out_path=None, pretags_path=None):
    out_path = out_path or DEFAULT_OUT
    pretags_path = pretags_path or PRETAGS
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    with open(pretags_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total = 0

    # ─── Sheet 1: 角色 (characters) ───
    def char_rows(ws, tb, lf):
        nonlocal total
        headers = ["序号", "cname", "source", "name", "appearance", "clothing",
                    "has_lora", "lora_file", "lora_hash", "lora_link",
                    "unet_weight", "clip_weight", "tags_count", "preview"]
        # rewrite headers row since make_ws already wrote it
        pass  
    
    # Actually let me just do a simpler approach
    for sheet_title, (data_key, col_headers, col_widths, field_map) in {
        "角色": ("characters", 
                 ["序号", "cname", "source", "name", "appearance", "clothing", "lora", 
                  "lora_file", "lora_hash", "lora_link", "unet_weight", "clip_weight", "tags_count"],
                 [6, 28, 14, 50, 60, 60, 6, 30, 16, 60, 12, 12, 10],
                 {1: "cname", 2: "source", 3: "name", 4: "appearance", 5: "clothing",
                  6: "has_lora", 7: "lora_file", 8: "lora_hash", 9: "lora_link",
                  10: "unet_weight", 11: "clip_weight", 12: "tags_count"}),
    }.items():
        ws = wb.create_sheet(title=sheet_title)
        hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hfn = Font(color="FFFFFF", bold=True, size=11)
        lf = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        tb = Border(left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9'))
        for col, h in enumerate(col_headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hf; c.font = hfn; c.alignment = Alignment(horizontal="center"); c.border = tb
        
        entries = data.get(data_key, {})
        row = 2
        for cname in sorted(entries.keys()):
            val = entries[cname]
            if not isinstance(val, dict):
                continue
            ws.cell(row=row, column=1, value=row-1).border = tb
            for col_idx, field_name in field_map.items():
                cell = ws.cell(row=row, column=col_idx+1, value=safe_str(val.get(field_name, "")))
                cell.border = tb
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            has_lora = val.get("has_lora", False)
            if has_lora:
                for c in range(1, len(col_headers) + 1):
                    ws.cell(row=row, column=c).fill = lf
            row += 1
        for i, w in enumerate(col_widths):
            ws.column_dimensions[col_letter(i + 1)].width = w
        ws.freeze_panes = "A2"
        last_row = row - 1
        if last_row > 1:
            ws.auto_filter.ref = f"A1:{col_letter(len(col_headers))}{last_row}"
        count = row - 2
        print(f"  {sheet_title}: {count} rows")
        total += count

    # ─── Sheets 2-7: categories (style/clothing/action/shot/scene/other) ───
    # 中文 sheet 标题 → 英文 data_key 映射
    _SHEET_KEY_MAP = {"画风": "style", "服装": "clothing", "动作": "action",
                      "镜头": "shot", "场景": "scene", "其他": "other"}
    for sheet_title in ["画风", "服装", "动作", "镜头", "场景", "其他"]:
        data_key = _SHEET_KEY_MAP[sheet_title]
        if data_key in ("style", "clothing"):
            fields = ["序号", "name", "tag", "has_lora", "lora_file", "lora_hash", "unet_weight", "clip_weight", "lora_link"]
            widths = [6, 36, 60, 6, 30, 16, 12, 12, 60]
        else:
            fields = ["序号", "name", "tag", "has_lora"]
            widths = [6, 36, 60, 6]
        
        ws = wb.create_sheet(title=sheet_title)
        hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hfn = Font(color="FFFFFF", bold=True, size=11)
        lf = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        tb = Border(left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9'))
        for col, h in enumerate(fields, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hf; c.font = hfn; c.alignment = Alignment(horizontal="center"); c.border = tb
        
        entries = data.get("categories", {}).get(data_key, {})
        row = 2
        for name in sorted(entries.keys()):
            val = entries[name]
            if not isinstance(val, dict):
                continue
            ws.cell(row=row, column=1, value=row-1).border = tb
            ws.cell(row=row, column=2, value=safe_str(val.get("name", ""))).border = tb
            cell = ws.cell(row=row, column=3, value=safe_str(val.get("tag", "")))
            cell.border = tb; cell.alignment = Alignment(wrap_text=True)
            has_lora = val.get("has_lora", False)
            ws.cell(row=row, column=4, value=1 if has_lora else 0).border = tb
            if data_key in ("style", "clothing") and has_lora:
                ws.cell(row=row, column=5, value=safe_str(val.get("lora_file", ""))).border = tb
                ws.cell(row=row, column=6, value=safe_str(val.get("lora_hash", ""))).border = tb
                ws.cell(row=row, column=7, value=safe_str(val.get("unet_weight", ""))).border = tb
                ws.cell(row=row, column=8, value=safe_str(val.get("clip_weight", ""))).border = tb
                ws.cell(row=row, column=9, value=safe_str(val.get("lora_link", ""))).border = tb
            if has_lora:
                for c in range(1, len(fields) + 1):
                    ws.cell(row=row, column=c).fill = lf
            row += 1
        
        for i, w in enumerate(widths):
            ws.column_dimensions[col_letter(i + 1)].width = w
        ws.freeze_panes = "A2"
        last_row = row - 1
        if last_row > 1:
            ws.auto_filter.ref = f"A1:{col_letter(len(fields))}{last_row}"
        count = row - 2
        print(f"  {sheet_title}: {count} rows")
        total += count

    # ─── Sheet 8: 系列 (series) ───
    sheet_title = "系列"
    fields = ["序号", "series_name", "count", "characters"]
    widths = [6, 30, 8, 120]
    ws = wb.create_sheet(title=sheet_title)
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfn = Font(color="FFFFFF", bold=True, size=11)
    tb = Border(left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9'))
    for col, h in enumerate(fields, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfn; c.alignment = Alignment(horizontal="center"); c.border = tb
    
    row = 2
    for sname in sorted(data.get("series", {}).keys()):
        sval = data["series"][sname]
        ws.cell(row=row, column=1, value=row-1).border = tb
        ws.cell(row=row, column=2, value=safe_str(sval.get("name", ""))).border = tb
        ws.cell(row=row, column=3, value=safe_str(sval.get("count", ""))).border = tb
        chars = sval.get("characters", [])
        cell = ws.cell(row=row, column=4, value=", ".join(chars) if chars else "")
        cell.border = tb; cell.alignment = Alignment(wrap_text=True)
        row += 1
    for i, w in enumerate(widths):
        ws.column_dimensions[col_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    last_row = row - 1
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{col_letter(len(fields))}{last_row}"
    count = row - 2
    print(f"  {sheet_title}: {count} rows")
    total += count

    wb.save(out_path)
    size_kb = os.path.getsize(out_path) / 1024
    print(f"\nTotal: {total} entries -> {out_path} ({size_kb:.0f} KB)")


if __name__ == "__main__":
    import argparse
    _p = argparse.ArgumentParser(description="Export pretags.json to multi-sheet Excel (current schema).")
    _p.add_argument("output", nargs="?", default=None, help="输出 xlsx 路径（默认 tmp/pretags_full_export.xlsx）")
    _p.add_argument("--pretags", "-p", default=None, help="pretags JSON 路径（默认项目根 pretags.json）")
    _a = _p.parse_args()
    main(out_path=_a.output, pretags_path=_a.pretags)
