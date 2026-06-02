#!/usr/bin/env python3
"""
Excel → Pretags 合并工具
用法: python pretags_merge_excel.py <excel_path> [--pretags <path>] [--dry-run]

按 cname 匹配：已有则更新，没有则新增。
操作前自动备份。
兼容旧格式（Source/Lora/model file name 等字段自动映射）。
"""
import json
import os
import shutil
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Try: pip install openpyxl")
    sys.exit(1)

# 使用相对路径
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
DEFAULT_PRETAGS = PROJECT_ROOT / 'pretags.json'
BACKUP_DIR = PROJECT_ROOT / 'Tanger-Presets-Show' / 'data' / 'backups'

FIELDS = ['cname', 'source', 'name', 'appearance', 'clothing',
          'has_lora', 'lora_file', 'lora_hash', 'lora_link',
          'unet_weight', 'clip_weight']

OLD_FIELD_MAP = {
    'Source': 'source', 'Lora': 'has_lora',
    'model file name': 'lora_file', 'model hash': 'lora_hash',
    'link': 'lora_link', '外貌': 'appearance', '服装': 'clothing',
}


def normalize_entry(entry):
    normalized = {}
    for k, v in entry.items():
        normalized[OLD_FIELD_MAP.get(k, k)] = v
    hl = normalized.get('has_lora')
    if hl == '是' or hl == 1:
        normalized['has_lora'] = True
    elif hl == '否' or hl == 0:
        normalized['has_lora'] = False
    return normalized


def build_tags(entry):
    """合并 name + appearance + clothing 为 tags 数组（2026-05-20 公子指定）"""
    tags = []
    for field in ['name', 'appearance', 'clothing']:
        val = str(entry.get(field, '')).strip().rstrip(',')
        if val:
            tags.extend([t.strip() for t in val.split(',') if t.strip()])
    return tags


def merge(excel_path, pretags_path, dry_run=False):
    wb = load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    excel_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            excel_data.append(normalize_entry(dict(zip(headers, row))))

    print(f"Excel: {len(excel_data)} 条")

    with open(pretags_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    chars = data.get('characters', {})
    added, updated, skipped = 0, 0, 0

    for entry in excel_data:
        cname = entry.get('cname', '')
        if not cname:
            skipped += 1
            continue
        tags = build_tags(entry)
        if cname not in chars:
            if not dry_run:
                chars[cname] = {f: entry.get(f, False if f == 'has_lora' else ([] if f == 'tags' else (0 if f == 'tags_count' else '')))
                                for f in FIELDS + ['tags', 'tags_count']}
                chars[cname]['tags'] = tags
                chars[cname]['tags_count'] = len(tags)
            added += 1
        else:
            if not dry_run:
                for f in FIELDS[1:]:
                    if f in entry and entry[f] is not None:
                        chars[cname][f] = entry[f]
                chars[cname]['tags'] = tags
                chars[cname]['tags_count'] = len(tags)
            updated += 1

    print(f"结果: 新增={added}, 更新={updated}, 跳过={skipped}, 总计={len(chars)}")
    if dry_run:
        print("[DRY RUN] 未写入"); return

    # 确保备份目录存在
    os.makedirs(BACKUP_DIR, exist_ok=True)
    
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = BACKUP_DIR / f'pretags_backup_{ts}.json'
    shutil.copy2(pretags_path, backup_path)
    with open(pretags_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已保存（备份: {backup_path}）")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('excel')
    parser.add_argument('--pretags', default=DEFAULT_PRETAGS)
    parser.add_argument('--dry-run', action='store_true')
    merge(**vars(parser.parse_args()))
