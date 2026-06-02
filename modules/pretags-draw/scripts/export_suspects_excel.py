#!/usr/bin/env python3
"""Export the 51 suspect parent entries to Excel for公子's review."""

import json, re, os, sys
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PRETAGS = os.path.join(os.path.dirname(__file__), "..", "..", "pretags.json")
OUT = os.path.join(os.path.dirname(__file__), "..", "tmp", "pretags_父条目审核.xlsx")

conflict_clothing = {'bikini', 'dress', 'swimsuit', 'skirt', 'bodysuit', 'robe', 
                     'armor', 'gown', 'suit', 'coat', 'jacket', 'kimono', 'yukata',
                     'lingerie', 'underwear', 'bra', 'panties', 'maid', 'school uniform',
                     'apron', 'cheongsam', 'qipao', 'nude', 'naked', 'towel'}

def col_letter(i):
    if i <= 26:
        return chr(64 + i)
    return chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)

def safe(v):
    if v is None:
        return ""
    return str(v)

def main():
    with open(PRETAGS, "r", encoding="utf-8") as f:
        data = json.load(f)
    chars = data['characters']

    # Group by lora_file
    by_lora = defaultdict(list)
    for cname, c in chars.items():
        lf = c.get('lora_file', '')
        if c.get('has_lora') and lf:
            by_lora[lf].append((cname, c))

    # Find suspects
    suspects = []
    for lf, entries in by_lora.items():
        if len(entries) <= 1:
            continue
        for cname, c in entries:
            tag_str = ' '.join(c.get('tags', [])).lower() + ' ' + c.get('appearance', '').lower()
            conflicts = {}
            for kw in conflict_clothing:
                cnt = tag_str.count(kw)
                if cnt > 0:
                    conflicts[kw] = cnt
            variants = len(re.findall(r'\b\w+[-_](?:def|s\d|outfit|costume|dress|armor|bikini|swim|suit)\b', tag_str))
            attr_counts = defaultdict(int)
            for tag in c.get('tags', []):
                for word in tag.lower().split():
                    if word in {'gloves', 'hair', 'skirt', 'dress', 'bikini', 'panties', 'bra', 'socks', 'thighhighs', 'boots', 'shoes'}:
                        attr_counts[word] += 1
            repeat_attrs = sum(1 for v in attr_counts.values() if v >= 3)
            tag_len = len(c.get('tags', []))
            
            is_parent = (
                (len(conflicts) >= 3 and sum(conflicts.values()) >= 6) or
                (len(conflicts) >= 2 and sum(conflicts.values()) >= 8) or
                variants >= 4 or
                repeat_attrs >= 2 or
                tag_len >= 20
            )
            if is_parent:
                siblings = sorted([e[0] for e in entries if e[0] != cname])
                suspects.append({
                    'cname': cname,
                    'source': c.get('source', ''),
                    'appearance': c.get('appearance', ''),
                    'clothing': c.get('clothing', ''),
                    'tags': c.get('tags', []),
                    'lora_file': c.get('lora_file', ''),
                    'lora_hash': c.get('lora_hash', ''),
                    'lora_link': c.get('lora_link', ''),
                    'unet_weight': c.get('unet_weight', ''),
                    'clip_weight': c.get('clip_weight', ''),
                    'preview': c.get('preview', ''),
                    'tags_count': tag_len,
                    'conflicts': conflicts,
                    'variants': variants,
                    'repeat_attrs': repeat_attrs,
                    'siblings': siblings,
                })

    suspects.sort(key=lambda x: -x['tags_count'])
    
    print(f"Suspects: {len(suspects)}")

    # ─── Write Excel ───
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "父条目审核"

    headers = ["序号", "cname", "source", "tags数", "冲突服装详情", "服装变体", "重复属性",
               "appearance", "clothing", "tags（完整）",
               "lora_file", "lora_hash", "lora_link", "unet", "clip", "preview", 
               "同lora子条目"]

    widths = [6, 36, 14, 8, 40, 10, 10, 80, 80, 100,
              30, 16, 60, 8, 8, 50, 60]

    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfn = Font(color="FFFFFF", bold=True, size=11)
    tb = Border(left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9'))
    lf = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfn; c.alignment = Alignment(horizontal="center"); c.border = tb

    row = 2
    for s in suspects:
        ws.cell(row=row, column=1, value=row-1).border = tb
        ws.cell(row=row, column=2, value=safe(s['cname'])).border = tb
        ws.cell(row=row, column=3, value=safe(s['source'])).border = tb
        ws.cell(row=row, column=4, value=s['tags_count']).border = tb
        conflicts_str = ', '.join(f"{k}×{v}" for k, v in sorted(s['conflicts'].items(), key=lambda x: -x[1]))
        ws.cell(row=row, column=5, value=conflicts_str).border = tb
        ws.cell(row=row, column=6, value=s['variants']).border = tb
        ws.cell(row=row, column=7, value=s['repeat_attrs']).border = tb

        c = ws.cell(row=row, column=8, value=safe(s['appearance']))
        c.border = tb; c.alignment = Alignment(wrap_text=True, vertical='top')
        
        c = ws.cell(row=row, column=9, value=safe(s['clothing']))
        c.border = tb; c.alignment = Alignment(wrap_text=True, vertical='top')
        
        tags_str = ', '.join(s['tags'])
        c = ws.cell(row=row, column=10, value=tags_str)
        c.border = tb; c.alignment = Alignment(wrap_text=True, vertical='top')

        ws.cell(row=row, column=11, value=safe(s['lora_file'])).border = tb
        ws.cell(row=row, column=12, value=safe(s['lora_hash'])).border = tb
        ws.cell(row=row, column=13, value=safe(s['lora_link'])).border = tb
        ws.cell(row=row, column=14, value=safe(s['unet_weight'])).border = tb
        ws.cell(row=row, column=15, value=safe(s['clip_weight'])).border = tb
        ws.cell(row=row, column=16, value=safe(s['preview'])).border = tb
        ws.cell(row=row, column=17, value=', '.join(s['siblings'])).border = tb

        # 高亮tags多的行
        if s['tags_count'] >= 20:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = lf

        row += 1

    for i, w in enumerate(widths):
        ws.column_dimensions[col_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    if row > 2:
        ws.auto_filter.ref = f"A1:{col_letter(len(headers))}{row-1}"

    wb.save(OUT)
    size_kb = os.path.getsize(OUT) / 1024
    print(f"✅ Saved: {OUT} ({size_kb:.0f} KB, {len(suspects)} rows)")

if __name__ == "__main__":
    main()
